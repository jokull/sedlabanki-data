import datetime as dt
import sqlite3
from io import BytesIO
from typing import List, Optional, Tuple

import arrow
import openpyxl
import requests
from openpyxl.worksheet.worksheet import Worksheet

from schemas import Row, Workbook, schemas

DBNAME = "credit.db"


def create_db(path=DBNAME):
    dbconn = sqlite3.connect(path, detect_types=sqlite3.PARSE_DECLTYPES)

    create_table_sql = """
    CREATE TABLE credit (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date DATE,
        institute VARCHAR,
        sector VARCHAR,
        industry VARCHAR,
        category VARCHAR,
        value INTEGER
    );
    """

    dbconn.cursor().execute(create_table_sql)

    return dbconn


def iter_series(wb_model: Workbook, wb: openpyxl.Workbook):
    for sheet_model in wb_model.sheets:
        sheet: Worksheet = wb.worksheets[sheet_model.sheet]  # type: ignore
        _series: List[Tuple[Row, List[Optional[int]]]] = []
        month_values = [
            c.value
            for c in sheet[sheet_model.dates_row][sheet_model.from_ : sheet.max_column]
        ]

        for row in sheet_model.rows:
            cells = sheet[row.row][sheet_model.from_ : sheet.max_column]
            assert len(month_values) == len(cells)
            values: List[Optional[int]] = [
                (int(c.value * 1_000_000) if c.value else None) for c in cells
            ]
            _series.append((row, values))

        yield (
            month_values,
            _series,
        )


def get_latest_workbook(wb_model: Workbook) -> Tuple[openpyxl.Workbook, str]:
    date = arrow.now()
    while True:
        url = wb_model.url.format(year=date.year, month=date.month)
        response = requests.get(url)
        if response.status_code == 404:
            date = date.shift(months=-1)
        else:
            break
    wb: openpyxl.Workbook = openpyxl.load_workbook(
        BytesIO(response.content), data_only=True
    )
    return wb, url


def main():

    dbconn = create_db()

    for wb_model in schemas:
        if wb_model.sheets is None:
            continue
        wb, url = get_latest_workbook(wb_model)

        print(url)

        for month_values, s in iter_series(wb_model, wb):
            for row, values in s:
                for date, value in zip(month_values, values):
                    dbconn.execute(
                        "INSERT INTO credit(institute, category, sector, industry, value, date) values (?, ?, ?, ?, ?, ?)",
                        (
                            row.institute,
                            row.category,
                            row.sector,
                            row.industry,
                            value,
                            date,
                        ),
                    )
            dbconn.commit()


if __name__ == "__main__":
    main()
